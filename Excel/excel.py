# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook

folder = 'C:\\Users\\Joseph\\Documents\\Programming\\Python\\Excel'

files  = ['Fondskickback.xlsx', 'Fondskosten.xlsx']


def open_workbook(fpath, fname):
    fwb = load_workbook(filename = os.path.join(folder, f))
    
    return fwb


def list_worksheets(ffwb):
    for ws in ffwb:
        print('ws ')
    

if __name__ == "__main__":
    if os.path.isdir(folder):
        for f in os.listdir(folder):
            if f.endswith('.xlsx'):
               wb = open_workbook(folder, f)
               print(str(wb.sheetnames))
